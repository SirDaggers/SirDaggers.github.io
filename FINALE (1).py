import numpy as np
import random
from tabulate import tabulate
import time
import os
from openpyxl import Workbook


# Constants
subjects = np.array(['Edukasyong Pantahanan at Pangkabuhayan', 'Edukasyon sa Pagpapakatao', 'Mathematics',
                     'BREAK', 'Filipino', 'Araling Panlipunan', 'English', 'MAPEH', 'Science'])
teachers = np.array(['Mrs. Abital', 'Mrs. Masigla', 'Mrs. Nadal', 'Mrs. Dulnuan', 'Mr. Paracha', 'Mrs. Fernandez'])
sections = np.array(["Daffodil", "Everlasting", "Sampaguita", "Daisy", "Camia"])
homeroom_teacher = np.array(["Mrs. Fevie Grace F. Habin", "Mrs. Julie Ann R. Abingona", "Mrs. Eglicelda A. Molo", "Mrs. Arlene A. Montes", "Mrs. Nery B. Sarion"])

# Define subject-time in minutes dictionary
subject_time = {
    'Edukasyon sa Pagpapakatao': 30,
    'Edukasyong Pantahanan at Pangkabuhayan': 40,
    'Science': 50,
    'Mathematics': 50,
    'BREAK': 10,
    'Filipino': 50,
    'Araling Panlipunan': 40,
    'English': 50,
    'MAPEH': 30,
}

# Define subject-teacher dictionary
subject_teacher = {
    'Edukasyon sa Pagpapakatao': 'Mrs. Abital',
    'Edukasyong Pantahanan at Pangkabuhayan': 'Mrs. Masigla',
    'Science': 'Mrs. Bello',
    'Mathematics': 'Mr. Tabuada',
    'BREAK': None,
    'Filipino': 'Mr. Paracha',
    'Araling Panlipunan': 'Mrs. Fernandez',
    'English': 'Mrs. Nadal',
    'MAPEH': 'Mrs. Dulnuan',
}

timeslots = 9

# Define constraints
no_homeroom_teacher_duplicates = True
no_teacher_duplicates = True
no_conflicting_timeslots = True
no_subject_duplicates = True
recess_timeslot = 4  # Timeslot index for RECESS

# Define the Particle class
class Particle:
    def __init__(self, num_sections):
        self.position = np.zeros((num_sections, timeslots), dtype=np.int64)
        self.best_position = np.zeros((num_sections, timeslots), dtype=np.int64)
        self.velocity = np.zeros((num_sections, timeslots), dtype=np.int64)
        self.fitness = float('inf')
        self.best_fitness = float('inf')

# Initialize the swarm
def initialize_swarm(num_particles, num_sections):
    swarm = []
    for _ in range(num_particles):
        particle = Particle(num_sections)
        swarm.append(particle)
    return swarm

# Generate a random schedule
def generate_schedule(num_sections):
    schedule = np.zeros((num_sections, timeslots), dtype=np.int64)
    for section_idx in range(num_sections):
        subjects_copy = np.copy(subjects)
        for timeslot in range(timeslots):
            if subjects_copy.size == 0:
                schedule[section_idx, timeslot] = np.where(subjects == 'BREAK')[0][0]  # Assign BREAK when subjects are exhausted
                continue
            subject_idx = random.choice(np.arange(subjects_copy.size))
            subject = subjects_copy[subject_idx]
            schedule[section_idx, timeslot] = np.where(subjects == subject)[0][0]
            subjects_copy = np.delete(subjects_copy, subject_idx)
    return schedule


def evaluate_fitness(schedule):
    penalties = []  # Array to store penalty values
    fitness = 0
    num_sections = schedule.shape[0]
    for section_idx in range(num_sections):
        section_schedule = schedule[section_idx]
        for timeslot in range(timeslots):
            subject_idx = section_schedule[timeslot]
            subject = subjects[subject_idx]
            if subject == 'BREAK':
                if timeslot != recess_timeslot:
                    penalties.append("BREAK")  # Store penalty value in the array
                    fitness += 100
                continue
            teacher = subject_teacher[subject]

            for other_timeslot in range(timeslots):
                if other_timeslot == timeslot:
                    continue
                other_subject_idx = section_schedule[other_timeslot]
                other_subject = subjects[other_subject_idx]
                if other_subject == 'BREAK':
                    continue
                other_teacher = subject_teacher[other_subject]
                if no_teacher_duplicates and teacher == other_teacher:
                    penalties.append("TeacherDups")  # Store penalty value in the array
                    fitness += 1000
                if no_subject_duplicates and subject == other_subject:
                    penalties.append("SubjectDups")  # Store penalty value in the array
                    fitness += 1000
                if no_conflicting_timeslots and other_timeslot == timeslot + subject_time[subject] - 1:
                    penalties.append("ConflictingTime")  # Store penalty value in the array
                    fitness += 1000
    return fitness, penalties

# Assign homeroom teachers to the schedule
def assign_homeroom_teachers(schedule):
    num_sections = schedule.shape[0]
    for section_idx in range(num_sections):
        homeroom = homeroom_teacher[section_idx]
        recess_subject_idx = np.where(subjects == 'BREAK')[0][0]
        recess_timeslot_indices = np.where(schedule[section_idx] == recess_subject_idx)[0]
        if recess_timeslot_indices.size > 0:
            selected_timeslot_idx = random.choice(recess_timeslot_indices)
            teacher_idx = np.where(teachers == homeroom)[0]
            if teacher_idx.size > 0:
                schedule[section_idx, selected_timeslot_idx] = teacher_idx[0]
    return schedule

# Perform the particle swarm optimization
def pso(num_particles, num_iterations, c1, c2, w, mutation_prob):
    swarm = initialize_swarm(num_particles, len(sections))
    global_best_fitness = float('inf')
    global_best_position = None
    for iteration in range(num_iterations):
        for particle in swarm:
            schedule = generate_schedule(len(sections))
            particle.position = schedule.copy()
            particle.fitness, penalties = evaluate_fitness(schedule)
            particle.position = assign_homeroom_teachers(particle.position)  # Assign homeroom teachers
            if particle.fitness < particle.best_fitness:
                particle.best_position = particle.position.copy()
                particle.best_fitness = particle.fitness
            if particle.fitness < global_best_fitness:
                global_best_fitness = particle.fitness
                global_best_position = particle.position.copy()

        for particle_idx, particle in enumerate(swarm):
            r1 = np.random.random(particle.velocity.shape) 
            r2 = np.random.random(particle.velocity.shape)
            particle.velocity = (w * particle.velocity +
                                 c1 * r1 * (particle.best_position - particle.position) +
                                 c2 * r2 * (global_best_position - particle.position))
            particle.velocity = np.clip(particle.velocity, -1, 1)
            particle.position = np.round(particle.position + particle.velocity).astype(np.int64)
            particle.position = np.clip(particle.position, 0, len(subjects) - 1)

            # Apply mutation
            if np.random.random() < mutation_prob:
                mutated_position = particle.position.copy()
                for section_idx in range(len(sections)):
                    if mutated_position[section_idx, 0] != 0:
                        # Swap the first subject (preliminary activities) with a random subject
                        rand_subject_idx = random.choice(np.arange(1, len(subjects)))
                        mutated_position[section_idx, 0], mutated_position[section_idx, rand_subject_idx] = \
                            mutated_position[section_idx, rand_subject_idx], mutated_position[section_idx, 0]

                mutated_fitness, penalties = evaluate_fitness(mutated_position)
                if mutated_fitness < particle.best_fitness:
                    particle.best_position = mutated_position.copy()
                    particle.best_fitness = mutated_fitness

        # Calculate and print the fitness for the global best position
        best_fitness, penalties = evaluate_fitness(global_best_position)
        print("Best Fitness (Iteration", iteration + 1, "):", best_fitness)

    return global_best_position

def clear_screen():
    # Clear the screen depending on the operating system
    os.system('cls' if os.name == 'nt' else 'clear')

def opening_design():
    print("\n\n===========================================================")
    print("Timetabling: Grade 5 Scheduling at Mambog Elementary School \n\t using Particle Swarm Optimization (PSO)")
    print("===========================================================")
    print("\n\t\tPress any key to continue")
    input()

    print("Loading...")
    for _ in range(10):
        print("█", end="", flush=True)
        time.sleep(0.2)

    clear_screen()

    print("====================================")
    print(" Welcome to the Timetabling System!")
    print("====================================\n")

def export_schedule_to_excel(best_schedule):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Write the best schedule to the worksheet
    num_sections = best_schedule.shape[0]
    row_idx = 1  # Start at row 1
    for section_idx in range(num_sections):
        section_name = sections[section_idx]
        homeroom = homeroom_teacher[section_idx]
        sheet.cell(row=row_idx, column=1).value = f"Section: {section_name} (Homeroom Teacher: {homeroom})"
        row_idx += 1  # Increment the row index

        # Write the headers for the schedule table
        headers = ["Time Starts", "Time Ends", "Minutes", "Subjects", "Teachers"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = header

        row_idx += 1  # Increment the row index

        table_data = []
        current_time = 6 * 60  # Start time at 6:00 AM in minutes
        for timeslot in range(timeslots):
            subject_idx = best_schedule[section_idx, timeslot]
            subject = subjects[subject_idx]
            subject_duration = subject_time[subject]
            time_start = f"{current_time // 60:02d}:{current_time % 60:02d}"
            current_time += subject_duration
            time_end = f"{current_time // 60:02d}:{current_time % 60:02d}"
            minutes = subject_duration
            if subject == 'BREAK':
                teacher = homeroom
            else:
                teacher = subject_teacher[subject]
            table_data.append([time_start, time_end, minutes, subject, teacher])

        for row_data in table_data:
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value
            row_idx += 1  # Increment the row index

        # Add an empty row between sections
        sheet.cell(row=row_idx, column=1).value = ""
        row_idx += 1  # Increment the row index

    # Save the workbook to a file
    workbook.save("schedule_results.xlsx")

# Call the opening_design function before the main function
opening_design()

# Run the scheduling system
best_schedule = pso(num_particles=20, num_iterations=100, c1=3.50, c2=0.50, w=0.85, mutation_prob=0.1)

# Calculate the fitness of the best schedule
best_fitness, penalties = evaluate_fitness(best_schedule)

# Export the schedule results to an Excel file
export_schedule_to_excel(best_schedule)

# Print a message indicating successful export
print("Schedule results exported to 'schedule_results.xlsx'.")
